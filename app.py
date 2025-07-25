import streamlit as st
import datetime
from collections import defaultdict
from openpyxl import Workbook
from fpdf import FPDF
import io

# Config
st.set_page_config(page_title="Gerador de Duplas", layout="centered")

# Dados
participants = [
    "Carlos Correia",
    "Jose Cocenas",
    "Antonio Ruas",
    "Jose Vendeiro",
    "Antonio Vidinha",
    "Arlindo Vendeiro"
]

availability = {
    "Antonio Vidinha": datetime.date(2025, 7, 25),
}

max_monthly = {
    "Arlindo Vendeiro": 1
}

fixed_participation = {
    datetime.date(2025, 7, 11): ("Carlos Correia", "Jose Vendeiro")
}

mandatory_days = {
    "Antonio Ruas": [
        datetime.date(2025, 7, 25),
        datetime.date(2025, 8, 29),
        datetime.date(2025, 9, 19),
        datetime.date(2025, 10, 17)
    ]
}

# Helpers
def get_fridays(start_date, end_date):
    current = start_date
    fridays = []
    while current <= end_date:
        if current.weekday() == 4:
            fridays.append(current)
        current += datetime.timedelta(days=1)
    return fridays

def format_date(date):
    return date.strftime("%d-%m-%Y")

def generate_schedule(start_month, end_month, year=2025):
    schedule = {}
    usage_tracker = defaultdict(list)
    arlindo_count = defaultdict(int)
    start_date = datetime.date(year, start_month, 1)
    end_date = datetime.date(year, end_month, 31)
    fridays = get_fridays(start_date, end_date)

    for date in fridays:
        if date in fixed_participation:
            p1, p2 = fixed_participation[date]
            schedule[date] = (p1, p2)
            usage_tracker[p1].append(date)
            usage_tracker[p2].append(date)
            continue

        must_have = None
        for name, days in mandatory_days.items():
            if date in days:
                must_have = name
                break

        possible = [p for p in participants
                    if availability.get(p, date) <= date
                    and (date not in usage_tracker[p][-1:] if usage_tracker[p] else True)]

        if must_have:
            possible = [p for p in possible if p != must_have]
            partner = next((p for p in possible if p != must_have and
                            not (p == "Arlindo Vendeiro" and arlindo_count[p] >= max_monthly.get(p, 99))), None)
            if partner:
                schedule[date] = (must_have, partner)
                usage_tracker[must_have].append(date)
                usage_tracker[partner].append(date)
                if "Arlindo Vendeiro" in (must_have, partner):
                    arlindo_count["Arlindo Vendeiro"] += 1
            else:
                schedule[date] = (must_have, "??")
        else:
            for i, p1 in enumerate(possible):
                for p2 in possible[i+1:]:
                    if "Arlindo Vendeiro" in (p1, p2):
                        if arlindo_count["Arlindo Vendeiro"] >= max_monthly.get("Arlindo Vendeiro", 99):
                            continue
                    last_friday = date - datetime.timedelta(days=7)
                    if last_friday in schedule and (p1 in schedule[last_friday] or p2 in schedule[last_friday]):
                        continue
                    schedule[date] = (p1, p2)
                    usage_tracker[p1].append(date)
                    usage_tracker[p2].append(date)
                    if "Arlindo Vendeiro" in (p1, p2):
                        arlindo_count["Arlindo Vendeiro"] += 1
                    break
                if date in schedule:
                    break
            if date not in schedule:
                schedule[date] = ("??", "??")

    return schedule

def export_excel(schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplas"
    ws.append(["Data", "Participante 1", "Participante 2"])
    for date in sorted(schedule):
        ws.append([format_date(date), *schedule[date]])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def export_pdf(schedule):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "Duplas por Sexta-feira", ln=True, align='C')
    pdf.ln(10)
    for date in sorted(schedule):
        p1, p2 = schedule[date]
        pdf.cell(0, 10, f"{format_date(date)}: {p1} & {p2}", ln=True)
    bio = io.BytesIO()
    pdf.output(bio)
    bio.seek(0)
    return bio

def export_txt(schedule):
    output = ""
    for date in sorted(schedule):
        p1, p2 = schedule[date]
        output += f"{format_date(date)}: {p1} & {p2}\n"
    return output.encode("utf-8")

# Interface
st.title("🎯 Gerador de Duplas por Sexta-feira")
col1, col2 = st.columns(2)
with col1:
    start_month = st.selectbox("📅 Mês de início", range(1, 13), index=6)
with col2:
    end_month = st.selectbox("📅 Mês de fim", range(1, 13), index=9)

if st.button("Gerar Duplas"):
    schedule = generate_schedule(start_month, end_month)
    st.success("✅ Duplas geradas com sucesso!\n")

    for date in sorted(schedule):
        p1, p2 = schedule[date]
        st.write(f"📌 **{format_date(date)}**: {p1} & {p2}")

    excel_data = export_excel(schedule)
    pdf_data = export_pdf(schedule)
    txt_data = export_txt(schedule)

    st.download_button("⬇️ Exportar Excel", excel_data, "duplas.xlsx")
    st.download_button("⬇️ Exportar PDF", pdf_data, "duplas.pdf")
    st.download_button("⬇️ Exportar Histórico (.txt)", txt_data, "duplas.txt")

